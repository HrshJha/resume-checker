"""
Excel formatter — converts a list of ``RankingExportRow`` into an
openpyxl ``Workbook`` with professional recruiter formatting.

Responsibilities:
    - Write header row with bold style and auto-filter.
    - Write data rows with correct cell types (int, float, str, datetime).
    - Apply percentage number format to score columns.
    - Apply a green-yellow-red color scale to the Final Score column.
    - Apply a blue-to-white color scale to Confidence column.
    - Freeze the first row (header).
    - Auto-size column widths using the schema-defined widths.
    - Apply alternating row shading for readability.
    - Apply conditional font coloring to the Recommendation column.

This module contains NO database logic and NO business logic.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from src.api.models.export_models import (
    COLUMN_SCHEMA,
    SCORE_COLUMNS,
    RankingExportRow,
)
from src.utils.logger import get_logger

logger = get_logger("excel_formatter")

# ---------------------------------------------------------------------------
# Style constants — defined once, referenced everywhere
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(
    start_color="1F3864",  # dark navy
    end_color="1F3864",
    fill_type="solid",
)
_HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=False
)

_DATA_FONT = Font(name="Calibri", size=10)
_DATA_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
_DATA_ALIGNMENT_LEFT = Alignment(
    horizontal="left", vertical="top", wrap_text=True
)

_ROW_FILL_ODD = PatternFill(
    start_color="F2F7FF",
    end_color="F2F7FF",
    fill_type="solid",
)
_ROW_FILL_EVEN = PatternFill(
    start_color="FFFFFF",
    end_color="FFFFFF",
    fill_type="solid",
)

_THIN_BORDER_SIDE = Side(border_style="thin", color="D0D7E0")
_CELL_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)

# Recommendation label → (font color hex, fill color hex)
_RECOMMENDATION_STYLES: dict[str, tuple[str, str]] = {
    "Strong Hire": ("155724", "D4EDDA"),  # dark green text, light green fill
    "Hire":        ("1C5A2E", "C3E6CB"),
    "Phone Screen":("856404", "FFF3CD"),  # amber
    "Pass":        ("721C24", "F8D7DA"),  # red
}

_PERCENTAGE_FORMAT = "0.00%"
_DATETIME_FORMAT = "YYYY-MM-DD HH:MM:SS"
_INTEGER_FORMAT = "0"
_FLOAT_FORMAT = "0.00"


# ---------------------------------------------------------------------------
# ExcelFormatter
# ---------------------------------------------------------------------------

class ExcelFormatter:
    """
    Formats a list of ``RankingExportRow`` objects as a styled XLSX workbook.

    Usage::

        formatter = ExcelFormatter()
        wb = formatter.format_workbook(rows, job_role="Senior Backend Engineer")
        wb.save("ranking.xlsx")
        # or:
        buf = formatter.format_to_bytes(rows)
    """

    def __init__(self) -> None:
        # Build column-index look-ups from the schema (1-based for openpyxl)
        self._col_names: list[str] = [col[0] for col in COLUMN_SCHEMA]
        self._col_headers: list[str] = [col[1] for col in COLUMN_SCHEMA]
        self._col_widths: list[int] = [col[2] for col in COLUMN_SCHEMA]
        self._n_cols: int = len(COLUMN_SCHEMA)

        # Map attribute name → 1-based column index
        self._col_index: dict[str, int] = {
            name: idx + 1 for idx, name in enumerate(self._col_names)
        }

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def format_workbook(
        self,
        rows: list[RankingExportRow],
        job_role: str = "Candidate Ranking",
        generated_at: datetime | None = None,
    ) -> Workbook:
        """
        Build and return a fully formatted openpyxl Workbook.

        Args:
            rows:         Ranked candidates (should already be sorted by rank).
            job_role:     Job title shown in the sheet tab name.
            generated_at: UTC timestamp for the export; defaults to now.

        Returns:
            An ``openpyxl.Workbook`` ready to save or stream.
        """
        if generated_at is None:
            from datetime import timezone as _tz
            generated_at = datetime.now(_tz.utc).replace(tzinfo=None)  # naive UTC for Excel

        wb = Workbook()

        # ── Main ranking sheet ──────────────────────────────────────────
        ws = wb.active
        sheet_title = f"Rankings - {job_role}"[:31]  # Excel limit: 31 chars
        ws.title = sheet_title  # type: ignore[union-attr]

        self._write_header(ws)           # type: ignore[arg-type]
        self._write_data_rows(ws, rows)  # type: ignore[arg-type]
        self._apply_column_widths(ws)    # type: ignore[arg-type]
        self._apply_freeze_and_filter(ws, n_data_rows=len(rows))  # type: ignore[arg-type]
        self._apply_conditional_formatting(ws, n_data_rows=len(rows))  # type: ignore[arg-type]

        # ── Summary sheet ───────────────────────────────────────────────
        ws_summary = wb.create_sheet("Export Summary")
        self._write_summary_sheet(ws_summary, rows, job_role, generated_at)

        logger.info(
            f"Workbook formatted: {len(rows)} rows, "
            f"{self._n_cols} columns, job_role='{job_role}'"
        )
        return wb

    def format_to_bytes(
        self,
        rows: list[RankingExportRow],
        job_role: str = "Candidate Ranking",
    ) -> bytes:
        """
        Format and return the workbook as raw bytes (for HTTP streaming).

        Args:
            rows:     Ranked candidate rows.
            job_role: Job title for sheet naming.

        Returns:
            Raw bytes of the XLSX file.
        """
        wb = self.format_workbook(rows, job_role=job_role)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ------------------------------------------------------------------
    # Private — header
    # ------------------------------------------------------------------

    def _write_header(self, ws: Any) -> None:
        """Write the bold, styled header row."""
        for col_idx, header in enumerate(self._col_headers, start=1):
            cell: Cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _CELL_BORDER

        ws.row_dimensions[1].height = 22  # type: ignore[attr-defined]

    # ------------------------------------------------------------------
    # Private — data rows
    # ------------------------------------------------------------------

    def _write_data_rows(self, ws: Any, rows: list[RankingExportRow]) -> None:
        """Write all data rows with correct types, formats, and fills."""
        for row_idx, export_row in enumerate(rows, start=2):
            fill = _ROW_FILL_ODD if row_idx % 2 == 0 else _ROW_FILL_EVEN

            for col_idx, attr_name in enumerate(self._col_names, start=1):
                raw_value = getattr(export_row, attr_name)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = _CELL_BORDER
                cell.fill = fill
                cell.font = _DATA_FONT

                self._write_cell(cell, attr_name, raw_value, export_row)

    def _write_cell(
        self,
        cell: Cell,
        attr_name: str,
        raw_value: Any,
        export_row: RankingExportRow,
    ) -> None:
        """Set cell value, number format, alignment, and special styles."""
        # ── Score columns ──────────────────────────────────────────────
        if attr_name in SCORE_COLUMNS:
            cell.value = float(raw_value) if raw_value is not None else 0.0
            cell.number_format = _PERCENTAGE_FORMAT
            cell.alignment = _DATA_ALIGNMENT_CENTER
            return

        # ── Rank (integer, centered) ───────────────────────────────────
        if attr_name == "rank":
            cell.value = int(raw_value)
            cell.number_format = _INTEGER_FORMAT
            cell.alignment = _DATA_ALIGNMENT_CENTER
            cell.font = Font(name="Calibri", size=10, bold=True)
            return

        # ── Years experience (1 decimal) ───────────────────────────────
        if attr_name == "years_experience":
            cell.value = float(raw_value) if raw_value is not None else 0.0
            cell.number_format = "0.0"
            cell.alignment = _DATA_ALIGNMENT_CENTER
            return

        # ── Timestamp ──────────────────────────────────────────────────
        if attr_name == "timestamp":
            if isinstance(raw_value, datetime):
                cell.value = raw_value.replace(tzinfo=None)  # Excel needs naive
            else:
                cell.value = str(raw_value)
            cell.number_format = _DATETIME_FORMAT
            cell.alignment = _DATA_ALIGNMENT_CENTER
            return

        # ── Recommendation (colored) ───────────────────────────────────
        if attr_name == "recommendation":
            cell.value = str(raw_value)
            style = _RECOMMENDATION_STYLES.get(str(raw_value))
            if style:
                font_color, fill_color = style
                cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=True,
                    color=font_color,
                )
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid",
                )
            cell.alignment = _DATA_ALIGNMENT_CENTER
            return

        # ── Long text columns (wrap) ───────────────────────────────────
        if attr_name in ("explanation", "matching_skills", "missing_skills",
                         "education", "resume_path"):
            cell.value = str(raw_value) if raw_value is not None else ""
            cell.alignment = _DATA_ALIGNMENT_LEFT
            return

        # ── ID columns (centered, monospace-ish) ──────────────────────
        if attr_name in ("candidate_id", "job_id"):
            cell.value = str(raw_value)
            cell.alignment = _DATA_ALIGNMENT_CENTER
            cell.font = Font(name="Courier New", size=9)
            return

        # ── Default ───────────────────────────────────────────────────
        cell.value = str(raw_value) if raw_value is not None else ""
        cell.alignment = _DATA_ALIGNMENT_CENTER

    # ------------------------------------------------------------------
    # Private — column widths
    # ------------------------------------------------------------------

    def _apply_column_widths(self, ws: Any) -> None:
        """Set column widths from the COLUMN_SCHEMA."""
        for col_idx, width in enumerate(self._col_widths, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width  # type: ignore[attr-defined]

        # Explanation and long text: let row height breathe
        for row_idx in range(2, ws.max_row + 1):  # type: ignore[attr-defined]
            ws.row_dimensions[row_idx].height = 60  # type: ignore[attr-defined]

    # ------------------------------------------------------------------
    # Private — freeze panes and auto-filter
    # ------------------------------------------------------------------

    def _apply_freeze_and_filter(self, ws: Any, n_data_rows: int) -> None:
        """
        Freeze the header row and apply an auto-filter to all columns.

        The auto-filter covers the entire data range so recruiters can
        filter by score, recommendation, etc. directly in Excel.
        """
        # Freeze row 1
        ws.freeze_panes = "A2"  # type: ignore[attr-defined]

        # Auto-filter on all columns
        last_col_letter = get_column_letter(self._n_cols)
        last_data_row = 1 + n_data_rows
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_data_row}"  # type: ignore[attr-defined]

    # ------------------------------------------------------------------
    # Private — conditional formatting
    # ------------------------------------------------------------------

    def _apply_conditional_formatting(
        self, ws: Any, n_data_rows: int
    ) -> None:
        """
        Apply Excel conditional formatting rules:

        1. Green-Yellow-Red 3-color scale on Final Score column.
        2. Blue-White color scale on Confidence column.
        3. Color-scale on Semantic Score, Career Score, Evidence Score.
        """
        if n_data_rows == 0:
            return

        first_data_row = 2
        last_data_row = 1 + n_data_rows

        def _range(col_name: str) -> str:
            col_letter = get_column_letter(self._col_index[col_name])
            return f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"

        # Final Score — green (good) to red (poor)
        ws.conditional_formatting.add(  # type: ignore[attr-defined]
            _range("final_score"),
            ColorScaleRule(
                start_type="num",   start_value=0.0, start_color="F8696B",
                mid_type="num",     mid_value=0.5,   mid_color="FFEB84",
                end_type="num",     end_value=1.0,   end_color="63BE7B",
            ),
        )

        # Semantic Score — white (no match) to dark blue (full match)
        ws.conditional_formatting.add(  # type: ignore[attr-defined]
            _range("semantic_score"),
            ColorScaleRule(
                start_type="min",   start_color="FFFFFF",
                end_type="max",     end_color="2F5597",
            ),
        )

        # Career Score — same green-yellow-red as final score
        ws.conditional_formatting.add(  # type: ignore[attr-defined]
            _range("career_score"),
            ColorScaleRule(
                start_type="num",   start_value=0.0, start_color="F8696B",
                mid_type="num",     mid_value=0.5,   mid_color="FFEB84",
                end_type="num",     end_value=1.0,   end_color="63BE7B",
            ),
        )

        # Evidence Score — white to teal
        ws.conditional_formatting.add(  # type: ignore[attr-defined]
            _range("evidence_score"),
            ColorScaleRule(
                start_type="min",   start_color="FFFFFF",
                end_type="max",     end_color="17A2B8",
            ),
        )

        # Confidence — light pink to dark green
        ws.conditional_formatting.add(  # type: ignore[attr-defined]
            _range("confidence"),
            ColorScaleRule(
                start_type="num",   start_value=0.0, start_color="FCE4EC",
                end_type="num",     end_value=1.0,   end_color="1B5E20",
            ),
        )

    # ------------------------------------------------------------------
    # Private — summary sheet
    # ------------------------------------------------------------------

    def _write_summary_sheet(
        self,
        ws: Any,
        rows: list[RankingExportRow],
        job_role: str,
        generated_at: datetime,
    ) -> None:
        """
        Write a compact summary sheet with export metadata and score
        distributions so the export is self-documenting.
        """
        h_font = Font(name="Calibri", bold=True, size=12)
        kv_font = Font(name="Calibri", size=10)

        def _h(row: int, value: str) -> None:
            cell = ws.cell(row=row, column=1, value=value)
            cell.font = h_font
            cell.fill = _HEADER_FILL
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=2
            )

        def _kv(row: int, key: str, value: Any) -> None:
            ws.cell(row=row, column=1, value=key).font = Font(
                name="Calibri", bold=True, size=10
            )
            ws.cell(row=row, column=2, value=value).font = kv_font

        ws.column_dimensions["A"].width = 30  # type: ignore[attr-defined]
        ws.column_dimensions["B"].width = 40  # type: ignore[attr-defined]

        r = 1
        _h(r, "Export Metadata")

        r += 1; _kv(r, "Job Role", job_role)
        r += 1; _kv(r, "Generated At (UTC)", generated_at.strftime("%Y-%m-%d %H:%M:%S"))
        r += 1; _kv(r, "Total Candidates Ranked", len(rows))

        if rows:
            scores = [row.final_score for row in rows]
            r += 1; _kv(r, "Top Score", f"{max(scores):.4f}")
            r += 1; _kv(r, "Median Score", f"{sorted(scores)[len(scores) // 2]:.4f}")
            r += 1; _kv(r, "Min Score", f"{min(scores):.4f}")

            from collections import Counter
            rec_counts = Counter(row.recommendation for row in rows)
            r += 1
            _h(r, "Recommendation Distribution")
            for label in ("Strong Hire", "Hire", "Phone Screen", "Pass"):
                r += 1
                _kv(r, label, rec_counts.get(label, 0))

        r += 1
        _h(r, "Score Column Reference")
        references = [
            ("Final Score",    "Weighted composite: 40% skills + 20% experience + 15% projects + 10% education + 5% BM25 + 5% CrossEncoder"),
            ("Semantic Score", "Required skill match ratio (matched required / total required)"),
            ("Evidence Score", "(has_projects + has_education) / 2"),
            ("Career Score",   "Experience years vs JD required years"),
            ("Behavior Score", "Normalised BM25 keyword relevance score"),
            ("Model Score",    "Same as Final Score (XGBoost LTR mirrors composite when model is untrained)"),
            ("Confidence",     "Derived from final_score: 0.10 + 0.80 * final_score, clamped to [0.10, 0.99]"),
        ]
        for ref_key, ref_desc in references:
            r += 1
            _kv(r, ref_key, ref_desc)
