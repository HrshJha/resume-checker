"""
Integration tests for the XLSX ranking export pipeline.

Tests the full path:
    register user → upload JD → upload candidates → rank → export via API

Verifies:
    - HTTP 200 with correct Content-Type
    - Content-Disposition header present with .xlsx filename
    - Workbook is valid and parseable by openpyxl
    - All required columns present
    - Rows sorted by rank ascending
    - No duplicate candidate IDs
    - Score values are floats in [0, 1]
    - Recommendation values are valid labels
    - HTTP 404 when JD not found
    - HTTP 409 when no rankings exist

These tests use an in-memory SQLite database (same pattern as the existing
integration tests in test_api_endpoints.py).
"""

from __future__ import annotations

import io

import openpyxl
import pytest
import pytest_asyncio
from httpx import AsyncClient, ASGITransport

from src.api.dependencies import get_settings
from src.api.main import app
from src.api.models.export_models import COLUMN_SCHEMA, SCORE_COLUMNS


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture
async def client(tmp_path):
    """Create a test ASGI client backed by a temporary SQLite database."""
    get_settings().db_url = f"sqlite+aiosqlite:///{tmp_path / 'test_export.db'}"
    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac


@pytest_asyncio.fixture
async def auth_headers(client: AsyncClient) -> dict[str, str]:
    """Register a test user and return bearer-token headers."""
    await client.post(
        "/api/v1/auth/register",
        json={
            "username": "export_tester",
            "password": "test_password_123",
            "role": "recruiter",
        },
    )
    login = await client.post(
        "/api/v1/auth/token",
        data={"username": "export_tester", "password": "test_password_123"},
    )
    token = login.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest_asyncio.fixture
async def jd_id(client: AsyncClient, auth_headers: dict) -> str:
    """Upload a sample JD and return its ID."""
    jd_text = (
        "Senior Python Engineer\n\n"
        "We need a Senior Python Engineer with 5+ years of experience.\n"
        "Requirements:\n"
        "- Python (5+ years)\n"
        "- FastAPI, PostgreSQL, Redis\n"
        "- Docker, Kubernetes\n"
        "- AWS or GCP\n\n"
        "Preferred:\n"
        "- Experience with ML pipelines\n"
        "- Open-source contributions\n\n"
        "Compensation: Competitive salary\n"
        "Location: Remote\n"
    )
    response = await client.post(
        "/api/v1/jobs/",
        json={"jd_text": jd_text},
        headers=auth_headers,
    )
    assert response.status_code == 201, f"JD upload failed: {response.text}"
    return response.json()["jd_id"]


@pytest_asyncio.fixture
async def ranked_jd_id(client: AsyncClient, auth_headers: dict, jd_id: str) -> str:
    """
    Upload minimal candidate resumes, wait for indexing, then run ranking.

    Returns the jd_id once rankings are persisted.

    Note: The test environment does not run real ML models. The ranking
    pipeline in search.py degrades gracefully to BM25 + heuristic scores
    when models are not loaded, so this fixture produces real (non-empty)
    rankings.
    """
    import anyio

    # Upload multiple small text-like files as resumes.
    # We create synthetic PDF-like bytes that the validator accepts.
    # The file validator checks magic bytes, so we craft minimal PDFs.
    minimal_pdf = (
        b"%PDF-1.4\n"
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
        b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
        b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] >>\nendobj\n"
        b"xref\n0 4\n0000000000 65535 f \n"
        b"trailer\n<< /Size 4 /Root 1 0 R >>\nstartxref\n0\n%%EOF"
    )

    # We need at least 1 successfully indexed candidate to test ranking.
    # Because real PDF parsing may fail in CI, we patch the processing path.
    # Instead we directly insert a candidate record via the DB so the ranking
    # pipeline has something to work with.
    from src.api.database import _async_session_factory
    from src.api.repositories.candidate_repo import CandidateRepository
    import uuid

    if _async_session_factory is None:
        pytest.skip("Database not initialized")

    candidate_ids = []
    async with _async_session_factory() as session:
        repo = CandidateRepository(session)
        for i in range(5):
            skills = ["Python", "FastAPI", "PostgreSQL", "Docker", "Redis", "AWS"]
            if i % 2 == 0:
                skills = skills[: 3 + i]  # Vary skill coverage
            cid = str(uuid.uuid4())
            candidate = await repo.create(
                candidate_id=cid,
                raw_resume_path=f"/data/uploads/{cid}.pdf",
                processing_status="indexed",
            )
            await session.execute(
                __import__("sqlalchemy", fromlist=["update"]).update(
                    __import__("src.api.db_models", fromlist=["Candidate"]).Candidate
                )
                .where(
                    __import__("src.api.db_models", fromlist=["Candidate"]).Candidate.candidate_id == cid
                )
                .values(
                    skills=skills,
                    experience_years=float(2 + i),
                    education=[{
                        "degree": "B.Tech",
                        "field_of_study": "Computer Science",
                        "institution": f"University {i}",
                        "end_year": str(2015 + i),
                    }],
                    projects=[{"name": "Project A", "description": "A project"}] if i % 2 == 0 else [],
                    parsed_data={
                        "full_name": f"Candidate {i + 1}",
                        "full_text": f"Candidate {i + 1}\nPython developer with {2 + i} years experience.",
                        "sections": {
                            "skills": " ".join(skills),
                            "experience": f"{2 + i} years of Python development",
                        },
                    },
                )
            )
            await session.commit()
            candidate_ids.append(cid)

    # Run ranking
    rank_resp = await client.post(
        "/api/v1/search/rank",
        json={"jd_id": jd_id, "top_k": 10},
        headers=auth_headers,
    )
    assert rank_resp.status_code == 200, f"Ranking failed: {rank_resp.text}"
    ranking_data = rank_resp.json()
    assert len(ranking_data["results"]) > 0, "No candidates ranked"

    return jd_id


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_export_returns_200(client, auth_headers, ranked_jd_id):
    """The export endpoint must return HTTP 200."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"


@pytest.mark.asyncio
async def test_export_content_type(client, auth_headers, ranked_jd_id):
    """Content-Type must be the OOXML MIME type."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    ct = response.headers.get("content-type", "")
    assert "spreadsheetml" in ct or "openxmlformats" in ct, (
        f"Unexpected Content-Type: {ct}"
    )


@pytest.mark.asyncio
async def test_export_content_disposition(client, auth_headers, ranked_jd_id):
    """Content-Disposition must signal attachment with .xlsx filename."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    cd = response.headers.get("content-disposition", "")
    assert "attachment" in cd, f"Missing 'attachment' in Content-Disposition: {cd}"
    assert ".xlsx" in cd, f"Missing .xlsx in Content-Disposition: {cd}"
    assert ranked_jd_id in cd, f"JD ID missing from filename: {cd}"


@pytest.mark.asyncio
async def test_export_is_valid_xlsx(client, auth_headers, ranked_jd_id):
    """Response body must be parseable as a valid XLSX workbook."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb is not None
    assert len(wb.sheetnames) >= 1


@pytest.mark.asyncio
async def test_export_all_columns_present(client, auth_headers, ranked_jd_id):
    """Every column defined in COLUMN_SCHEMA must be present in the XLSX."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None

    expected_headers = [col[1] for col in COLUMN_SCHEMA]
    actual_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMN_SCHEMA))]
    assert actual_headers == expected_headers, (
        f"Header mismatch.\nExpected: {expected_headers}\nActual: {actual_headers}"
    )


@pytest.mark.asyncio
async def test_export_rows_sorted_by_rank(client, auth_headers, ranked_jd_id):
    """Rows must be sorted ascending by the Rank column."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None

    ranks = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None:
            ranks.append(val)

    assert ranks == sorted(ranks), f"Ranks not sorted: {ranks}"


@pytest.mark.asyncio
async def test_export_no_duplicate_candidate_ids(client, auth_headers, ranked_jd_id):
    """No candidate_id should appear more than once in the export."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None

    # Find candidate_id column index
    cid_col = next(
        i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "candidate_id"
    )
    cids = [
        ws.cell(row=i, column=cid_col).value
        for i in range(2, ws.max_row + 1)
        if ws.cell(row=i, column=cid_col).value
    ]
    assert len(cids) == len(set(cids)), f"Duplicate candidate IDs: {cids}"


@pytest.mark.asyncio
async def test_export_scores_are_floats_in_range(client, auth_headers, ranked_jd_id):
    """All score column values must be floats in [0, 1]."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None

    score_col_indices = {
        name: i + 1
        for i, (name, _, _) in enumerate(COLUMN_SCHEMA)
        if name in SCORE_COLUMNS
    }

    for col_name, col_idx in score_col_indices.items():
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            assert isinstance(val, float), (
                f"Column '{col_name}' row {row_idx}: expected float, got {type(val)}"
            )
            assert 0.0 <= val <= 1.0, (
                f"Column '{col_name}' row {row_idx}: value {val} out of [0, 1]"
            )


@pytest.mark.asyncio
async def test_export_recommendations_are_valid(client, auth_headers, ranked_jd_id):
    """Recommendation column must contain only valid label values."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None

    valid_labels = {"Strong Hire", "Hire", "Phone Screen", "Pass"}
    rec_col = next(
        i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "recommendation"
    )
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=rec_col).value
        if val is not None:
            assert val in valid_labels, (
                f"Invalid recommendation '{val}' at row {row_idx}"
            )


@pytest.mark.asyncio
async def test_export_scores_match_rank_api(client, auth_headers, ranked_jd_id):
    """Final scores in the XLSX must match scores returned by the rank API."""
    # Get rankings from API
    rank_resp = await client.post(
        "/api/v1/search/rank",
        json={"jd_id": ranked_jd_id, "top_k": 10},
        headers=auth_headers,
    )
    api_scores = {
        r["candidate_id"]: r["final_score"]
        for r in rank_resp.json()["results"]
    }

    # Get XLSX export
    export_resp = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
    ws = wb.active
    assert ws is not None

    cid_col = next(
        i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "candidate_id"
    )
    score_col = next(
        i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "final_score"
    )

    for row_idx in range(2, ws.max_row + 1):
        cid = ws.cell(row=row_idx, column=cid_col).value
        xlsx_score = ws.cell(row=row_idx, column=score_col).value
        if cid and cid in api_scores:
            api_score = api_scores[cid]
            # Allow small floating-point delta (scores are rounded to 4 or 6 dp)
            assert abs(float(xlsx_score) - float(api_score)) < 1e-4, (
                f"Score mismatch for {cid}: XLSX={xlsx_score}, API={api_score}"
            )


@pytest.mark.asyncio
async def test_export_returns_404_for_missing_jd(client, auth_headers):
    """Export must return 404 when the JD does not exist."""
    response = await client.get(
        "/api/v1/search/rank/nonexistent-jd-id-00000/export",
        headers=auth_headers,
    )
    assert response.status_code == 404


@pytest.mark.asyncio
async def test_export_returns_409_when_no_rankings(client, auth_headers, jd_id):
    """Export must return 409 when the JD exists but has no rankings."""
    # jd_id fixture creates a JD but does NOT run ranking
    response = await client.get(
        f"/api/v1/search/rank/{jd_id}/export",
        headers=auth_headers,
    )
    assert response.status_code == 409


@pytest.mark.asyncio
async def test_export_requires_authentication(client, ranked_jd_id):
    """Export endpoint must require a valid JWT."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export"
        # No auth headers
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_export_top_k_respected(client, auth_headers, ranked_jd_id):
    """Setting top_k=2 must yield at most 2 data rows."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export?top_k=2",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None
    n_data_rows = ws.max_row - 1  # subtract header
    assert n_data_rows <= 2, f"Expected <=2 data rows, got {n_data_rows}"


@pytest.mark.asyncio
async def test_workbook_has_summary_sheet(client, auth_headers, ranked_jd_id):
    """The exported workbook must contain an 'Export Summary' sheet."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Export Summary" in wb.sheetnames


@pytest.mark.asyncio
async def test_frozen_pane_in_exported_file(client, auth_headers, ranked_jd_id):
    """The exported file must have the header row frozen."""
    response = await client.get(
        f"/api/v1/search/rank/{ranked_jd_id}/export",
        headers=auth_headers,
    )
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws is not None
    assert ws.freeze_panes == "A2"
