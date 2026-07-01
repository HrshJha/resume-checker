"""
Unit tests for ExcelFormatter, export_models helpers, and ExportService
pure-function helpers.

These tests run entirely in-memory — no database, no HTTP server.
They verify:
    - COLUMN_SCHEMA completeness and field ordering
    - score_to_recommendation thresholds
    - score_to_confidence monotonicity
    - ExcelFormatter column count, header names, row ordering
    - ExcelFormatter score formatting (percentage cells)
    - ExcelFormatter frozen panes and auto-filter
    - ExcelFormatter no duplicate ranks
    - ExcelFormatter conditional formatting presence
    - _extract_candidate_name fallback chain
    - _extract_education_summary formatting
    - _build_inline_explanation content
"""

from __future__ import annotations

import io
import re
from dataclasses import fields as dc_fields
from datetime import datetime, timezone
from types import SimpleNamespace
from typing import Any
from unittest.mock import MagicMock

import openpyxl
import pytest

from src.api.models.export_models import (
    COLUMN_SCHEMA,
    RECOMMENDATION_THRESHOLDS,
    SCORE_COLUMNS,
    RankingExportRow,
    score_to_confidence,
    score_to_recommendation,
)
from src.api.services.excel_formatter import ExcelFormatter
from src.api.services.export_service import (
    _build_inline_explanation,
    _extract_candidate_name,
    _extract_education_summary,
    _extract_skill_lists,
    _format_skills_list,
)


# ===========================================================================
# Fixtures
# ===========================================================================

def _make_row(rank: int = 1, final_score: float = 0.75, **kwargs) -> RankingExportRow:
    """Create a minimal valid RankingExportRow for testing."""
    defaults: dict[str, Any] = dict(
        rank=rank,
        candidate_id=f"cand-{rank:04d}",
        candidate_name=f"Candidate {rank}",
        job_id="jd-0001",
        resume_path=f"/data/uploads/cand-{rank:04d}.pdf",
        final_score=final_score,
        semantic_score=0.80,
        evidence_score=0.60,
        career_score=0.90,
        behavior_score=0.55,
        model_score=final_score,
        confidence=score_to_confidence(final_score),
        matching_skills="Python, Docker, PostgreSQL",
        missing_skills="Go, Kubernetes",
        years_experience=5.5,
        education="B.Tech, Computer Science, IIT Delhi (2018)",
        recommendation=score_to_recommendation(final_score),
        explanation="Candidate is well-matched.",
        timestamp=datetime(2026, 7, 1, 12, 0, 0, tzinfo=timezone.utc),
    )
    defaults.update(kwargs)
    return RankingExportRow(**defaults)


def _make_rows(n: int = 5) -> list[RankingExportRow]:
    """Return n rows with descending final scores."""
    return [
        _make_row(rank=i, final_score=round(0.95 - (i - 1) * 0.07, 4))
        for i in range(1, n + 1)
    ]


@pytest.fixture
def formatter() -> ExcelFormatter:
    return ExcelFormatter()


@pytest.fixture
def sample_rows() -> list[RankingExportRow]:
    return _make_rows(10)


@pytest.fixture
def workbook_from_rows(formatter, sample_rows) -> openpyxl.Workbook:
    return formatter.format_workbook(sample_rows, job_role="Senior Backend Engineer")


@pytest.fixture
def ws(workbook_from_rows):
    return workbook_from_rows.active


# ===========================================================================
# Column schema tests
# ===========================================================================

class TestColumnSchema:
    def test_schema_length_matches_dataclass_fields(self):
        """Every COLUMN_SCHEMA entry must correspond to a dataclass field."""
        field_names = {f.name for f in dc_fields(RankingExportRow)}
        schema_names = {col[0] for col in COLUMN_SCHEMA}
        assert schema_names == field_names

    def test_no_duplicate_attribute_names(self):
        seen: set[str] = set()
        for attr_name, _, _ in COLUMN_SCHEMA:
            assert attr_name not in seen, f"Duplicate column: {attr_name}"
            seen.add(attr_name)

    def test_no_duplicate_display_headers(self):
        seen: set[str] = set()
        for _, header, _ in COLUMN_SCHEMA:
            assert header not in seen, f"Duplicate header: {header}"
            seen.add(header)

    def test_all_column_widths_positive(self):
        for attr_name, _, width in COLUMN_SCHEMA:
            assert width > 0, f"Width <= 0 for column '{attr_name}'"

    def test_score_columns_are_subset_of_schema(self):
        schema_attrs = {col[0] for col in COLUMN_SCHEMA}
        for score_col in SCORE_COLUMNS:
            assert score_col in schema_attrs, f"Score column '{score_col}' not in schema"


# ===========================================================================
# score_to_recommendation tests
# ===========================================================================

class TestScoreToRecommendation:
    def test_strong_hire_at_0_85(self):
        assert score_to_recommendation(0.85) == "Strong Hire"

    def test_strong_hire_at_0_80(self):
        assert score_to_recommendation(0.80) == "Strong Hire"

    def test_hire_at_0_70(self):
        assert score_to_recommendation(0.70) == "Hire"

    def test_hire_at_0_65(self):
        assert score_to_recommendation(0.65) == "Hire"

    def test_phone_screen_at_0_55(self):
        assert score_to_recommendation(0.55) == "Phone Screen"

    def test_phone_screen_at_0_50(self):
        assert score_to_recommendation(0.50) == "Phone Screen"

    def test_pass_at_0_49(self):
        assert score_to_recommendation(0.49) == "Pass"

    def test_pass_at_0_0(self):
        assert score_to_recommendation(0.0) == "Pass"

    def test_thresholds_cover_full_range(self):
        """score_to_recommendation must return a non-empty string for all values."""
        for v in [0.0, 0.1, 0.5, 0.64, 0.65, 0.79, 0.80, 1.0]:
            result = score_to_recommendation(v)
            assert isinstance(result, str) and result, f"Empty result at {v}"


# ===========================================================================
# score_to_confidence tests
# ===========================================================================

class TestScoreToConfidence:
    def test_confidence_monotonic(self):
        """Confidence must be non-decreasing as score increases."""
        scores = [round(i * 0.05, 2) for i in range(21)]  # 0.00 to 1.00
        confidences = [score_to_confidence(s) for s in scores]
        for i in range(1, len(confidences)):
            assert confidences[i] >= confidences[i - 1], (
                f"Confidence not monotonic at score={scores[i]}"
            )

    def test_confidence_clamped_below(self):
        assert score_to_confidence(0.0) >= 0.10

    def test_confidence_clamped_above(self):
        assert score_to_confidence(1.0) <= 0.99

    def test_confidence_in_range(self):
        for v in [0.0, 0.25, 0.5, 0.75, 1.0]:
            c = score_to_confidence(v)
            assert 0.0 <= c <= 1.0, f"Confidence out of range at score={v}"

    def test_confidence_precision(self):
        """Confidence should have at most 4 decimal places."""
        c = score_to_confidence(0.567)
        assert c == round(c, 4)


# ===========================================================================
# ExcelFormatter — structure tests
# ===========================================================================

class TestExcelFormatterStructure:
    def test_returns_workbook(self, formatter, sample_rows):
        wb = formatter.format_workbook(sample_rows)
        assert isinstance(wb, openpyxl.Workbook)

    def test_active_sheet_exists(self, workbook_from_rows):
        assert workbook_from_rows.active is not None

    def test_sheet_count(self, workbook_from_rows):
        """Should have exactly 2 sheets: Rankings + Export Summary."""
        assert len(workbook_from_rows.sheetnames) == 2

    def test_summary_sheet_present(self, workbook_from_rows):
        assert "Export Summary" in workbook_from_rows.sheetnames

    def test_correct_column_count(self, ws):
        assert ws.max_column == len(COLUMN_SCHEMA)

    def test_header_row_values(self, ws):
        expected_headers = [col[1] for col in COLUMN_SCHEMA]
        actual_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMN_SCHEMA))]
        assert actual_headers == expected_headers

    def test_data_row_count(self, ws, sample_rows):
        """Data rows = total rows - 1 header row."""
        assert ws.max_row == len(sample_rows) + 1

    def test_no_empty_header_cell(self, ws):
        for col_idx in range(1, len(COLUMN_SCHEMA) + 1):
            val = ws.cell(row=1, column=col_idx).value
            assert val, f"Empty header at column {col_idx}"


# ===========================================================================
# ExcelFormatter — cell content tests
# ===========================================================================

class TestExcelFormatterCellContent:
    def test_rank_values_correct(self, ws, sample_rows):
        for i, row in enumerate(sample_rows, start=2):
            cell_val = ws.cell(row=i, column=1).value
            assert cell_val == row.rank, f"Rank mismatch at row {i}"

    def test_no_duplicate_ranks(self, ws, sample_rows):
        ranks = [ws.cell(row=i, column=1).value for i in range(2, len(sample_rows) + 2)]
        assert len(ranks) == len(set(ranks)), "Duplicate ranks in output"

    def test_final_score_cell_is_float(self, ws):
        # Column index of final_score
        col_idx = next(
            i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "final_score"
        )
        cell_val = ws.cell(row=2, column=col_idx).value
        assert isinstance(cell_val, float), f"final_score should be float, got {type(cell_val)}"

    def test_final_score_in_range(self, ws, sample_rows):
        col_idx = next(
            i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "final_score"
        )
        for row_idx in range(2, len(sample_rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            assert 0.0 <= val <= 1.0, f"Score out of [0,1] range at row {row_idx}"

    def test_score_columns_have_percentage_format(self, ws):
        score_col_indices = {
            i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name in SCORE_COLUMNS
        }
        for col_idx in score_col_indices:
            cell = ws.cell(row=2, column=col_idx)
            assert "%" in (cell.number_format or ""), (
                f"Column {col_idx} missing percentage format: '{cell.number_format}'"
            )

    def test_candidate_id_is_string(self, ws):
        col_idx = next(
            i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "candidate_id"
        )
        val = ws.cell(row=2, column=col_idx).value
        assert isinstance(val, str)

    def test_timestamp_is_datetime(self, ws):
        col_idx = next(
            i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "timestamp"
        )
        val = ws.cell(row=2, column=col_idx).value
        assert isinstance(val, datetime)

    def test_rows_sorted_by_rank(self, ws, sample_rows):
        n = len(sample_rows)
        ranks = [ws.cell(row=i, column=1).value for i in range(2, n + 2)]
        assert ranks == sorted(ranks)

    def test_recommendation_cells_have_value(self, ws, sample_rows):
        col_idx = next(
            i + 1 for i, (name, _, _) in enumerate(COLUMN_SCHEMA) if name == "recommendation"
        )
        valid = {"Strong Hire", "Hire", "Phone Screen", "Pass"}
        for row_idx in range(2, len(sample_rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            assert val in valid, f"Unexpected recommendation '{val}' at row {row_idx}"


# ===========================================================================
# ExcelFormatter — format_to_bytes tests
# ===========================================================================

class TestExcelFormatterToBytes:
    def test_returns_bytes(self, formatter, sample_rows):
        raw = formatter.format_to_bytes(sample_rows)
        assert isinstance(raw, bytes)
        assert len(raw) > 0

    def test_bytes_parseable_as_xlsx(self, formatter, sample_rows):
        raw = formatter.format_to_bytes(sample_rows)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert len(wb.sheetnames) >= 1

    def test_bytes_roundtrip_columns(self, formatter, sample_rows):
        raw = formatter.format_to_bytes(sample_rows)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        assert ws.max_column == len(COLUMN_SCHEMA)  # type: ignore[union-attr]

    def test_bytes_roundtrip_row_count(self, formatter, sample_rows):
        raw = formatter.format_to_bytes(sample_rows)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        # max_row = header + data rows
        assert ws.max_row == len(sample_rows) + 1  # type: ignore[union-attr]

    def test_empty_rows_produces_header_only(self, formatter):
        raw = formatter.format_to_bytes([])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        assert ws.max_row == 1  # type: ignore[union-attr]
        assert ws.max_column == len(COLUMN_SCHEMA)  # type: ignore[union-attr]


# ===========================================================================
# ExcelFormatter — frozen panes and auto-filter
# ===========================================================================

class TestExcelFormatterPanesAndFilter:
    def test_frozen_pane_at_A2(self, ws):
        assert ws.freeze_panes == "A2"

    def test_auto_filter_set(self, ws):
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith("A1")

    def test_auto_filter_covers_all_columns(self, ws):
        ref = ws.auto_filter.ref or ""
        # Extract the end column letter from the ref (e.g. "A1:S11" → "S")
        match = re.search(r":([A-Z]+)\d+", ref)
        assert match, f"Could not parse auto_filter ref: {ref}"
        from openpyxl.utils import column_index_from_string
        end_col_idx = column_index_from_string(match.group(1))
        assert end_col_idx == len(COLUMN_SCHEMA)


# ===========================================================================
# ExcelFormatter — conditional formatting presence
# ===========================================================================

class TestConditionalFormatting:
    def test_conditional_formatting_exists(self, ws):
        cf_rules = ws.conditional_formatting
        assert len(cf_rules) > 0, "No conditional formatting rules applied"

    def test_final_score_has_color_scale(self, ws):
        """
        Verify that at least one conditional formatting rule uses a color scale.

        openpyxl stores CF rules as openpyxl.formatting.rule.Rule objects
        (not as ColorScaleRule instances directly).  A color scale rule has a
        non-None ``colorScale`` attribute on the Rule object.
        """
        has_color_scale = False
        for _, rules in ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                # openpyxl wraps ColorScaleRule inside a Rule.colorScale attr
                if getattr(rule, "colorScale", None) is not None:
                    has_color_scale = True
                    break
        assert has_color_scale, "No color-scale conditional formatting rule found on any column"


# ===========================================================================
# ExportService pure-function helpers
# ===========================================================================

class TestExtractCandidateName:
    def _make_candidate(self, parsed_data: dict) -> Any:
        c = MagicMock()
        c.parsed_data = parsed_data
        c.candidate_id = "test-uuid-1234"
        return c

    def test_returns_full_name_key(self):
        c = self._make_candidate({"full_name": "Jane Smith"})
        assert _extract_candidate_name(c) == "Jane Smith"

    def test_returns_name_key(self):
        c = self._make_candidate({"name": "John Doe"})
        assert _extract_candidate_name(c) == "John Doe"

    def test_prefers_full_name_over_name(self):
        c = self._make_candidate({"full_name": "Alice", "name": "Bob"})
        assert _extract_candidate_name(c) == "Alice"

    def test_falls_back_to_first_text_line(self):
        c = self._make_candidate({"full_text": "Robert Johnson\nsome@email.com\n..."})
        name = _extract_candidate_name(c)
        assert name == "Robert Johnson"

    def test_skips_email_lines(self):
        c = self._make_candidate({"full_text": "some@email.com\nCarol Williams\n..."})
        name = _extract_candidate_name(c)
        assert name == "Carol Williams"

    def test_fallback_to_candidate_id(self):
        c = self._make_candidate({"full_text": ""})
        name = _extract_candidate_name(c)
        assert name == "test-uuid-1234"

    def test_empty_parsed_data(self):
        c = self._make_candidate({})
        name = _extract_candidate_name(c)
        assert name == "test-uuid-1234"


class TestExtractEducationSummary:
    def _make_candidate(self, education: list) -> Any:
        c = MagicMock()
        c.education = education
        return c

    def test_returns_empty_for_no_education(self):
        c = self._make_candidate([])
        assert _extract_education_summary(c) == ""

    def test_formats_full_entry(self):
        c = self._make_candidate([{
            "degree": "B.Tech",
            "field_of_study": "Computer Science",
            "institution": "IIT Delhi",
            "end_year": "2018",
        }])
        result = _extract_education_summary(c)
        assert "B.Tech" in result
        assert "Computer Science" in result
        assert "IIT Delhi" in result
        assert "2018" in result

    def test_handles_missing_fields_gracefully(self):
        c = self._make_candidate([{"degree": "MBA"}])
        result = _extract_education_summary(c)
        assert "MBA" in result

    def test_returns_string(self):
        c = self._make_candidate([{"institution": "BITS Pilani"}])
        assert isinstance(_extract_education_summary(c), str)


class TestFormatSkillsList:
    def test_empty(self):
        assert _format_skills_list([]) == ""

    def test_single(self):
        assert _format_skills_list(["Python"]) == "Python"

    def test_multiple(self):
        result = _format_skills_list(["Python", "Go", "Docker"])
        assert result == "Python, Go, Docker"

    def test_truncates_at_200(self):
        long_list = ["skill_" + str(i) for i in range(100)]
        result = _format_skills_list(long_list)
        assert len(result) <= 200

    def test_skips_none_values(self):
        result = _format_skills_list([None, "Python", None])  # type: ignore[list-item]
        assert "Python" in result
        assert "None" not in result


class TestBuildInlineExplanation:
    def _make_ranking(self, **kwargs) -> Any:
        r = MagicMock()
        r.rank = kwargs.get("rank", 1)
        r.semantic_score = kwargs.get("semantic_score", 0.80)
        r.career_score = kwargs.get("career_score", 0.90)
        r.behavior_score = kwargs.get("behavior_score", 0.50)
        r.evidence_score = kwargs.get("evidence_score", 0.70)
        r.final_score = kwargs.get("final_score", 0.75)
        return r

    def test_contains_rank(self):
        r = self._make_ranking(rank=3)
        explanation = _build_inline_explanation(r)
        assert "#3" in explanation

    def test_contains_final_score_percentage(self):
        r = self._make_ranking(final_score=0.75)
        explanation = _build_inline_explanation(r)
        assert "75.0%" in explanation

    def test_contains_recommendation(self):
        r = self._make_ranking(final_score=0.85)
        explanation = _build_inline_explanation(r)
        assert "Strong Hire" in explanation

    def test_returns_string(self):
        r = self._make_ranking()
        assert isinstance(_build_inline_explanation(r), str)
